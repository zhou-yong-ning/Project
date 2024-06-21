import openpyxl
import os
import pandas as pd


def xlsx_value(file, str_txt_list):
    # 读取文件
    workbook = openpyxl.load_workbook(file)
    # 指定当前工作表
    worksheet = workbook.active
    # 写入单元格值
    worksheet.cell(16, 2, value=str_txt_list[6])
    worksheet.cell(17, 2, value=str_txt_list[7])
    worksheet.cell(18, 2, value=str_txt_list[4])
    worksheet.cell(19, 2, value=str_txt_list[5])
    worksheet.cell(20, 2, value=str_txt_list[0])
    worksheet.cell(21, 2, value=str_txt_list[1])
    worksheet.cell(22, 2, value=str_txt_list[2])
    worksheet.cell(23, 2, value=str_txt_list[3])
    worksheet.cell(4, 2, value=str_txt_list[8])
    # 保存
    workbook.save(os.path.join('NewFolder', str_txt_list[-1]+'.xlsx'))


Current_Folder_path = os.getcwd()
if not os.path.exists(Current_Folder_path + '\\NewFolder'):
    os.mkdir(Current_Folder_path + '\\NewFolder')
df1 = pd.read_excel('团风元数据对应表.xlsx', dtype='object')
for i in df1.index.tolist():
    # pandas条件查询
    dnx = df1.loc[i, '东南X']
    dny = df1.loc[i, '东南Y']
    dbx = df1.loc[i, '东北X']
    dby = df1.loc[i, '东北Y']
    xnx = df1.loc[i, '西南X']
    xny = df1.loc[i, '西南Y']
    xbx = df1.loc[i, '西北X']
    xby = df1.loc[i, '西北Y']
    name = df1.loc[i, '图号']
    xlsx_value('模板.xlsx', [dbx, dby, dnx, dny, xbx, xby, xnx, xny, name])
    print(name)
