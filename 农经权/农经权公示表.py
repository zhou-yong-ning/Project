import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


def range_sum(worksheet, start, end):
    sum = 0
    for row in worksheet[start:end]:
        for cell in row:
            if cell.value != None:
                sum += cell.value
    return sum


def is_blank_row(worksheet, row_num, max_col=None):
    if not max_col:
        max_col = worksheet.max_column
    for cell in worksheet[row_num][:max_col]:
        if cell.value:
            return False
    return True


def total_amount(worksheet):
    """ 对某sheet的A、E列合并居中，并对E列求和 """
    ws = worksheet
    row, max_row = 2, ws.max_row
    while row < ws.max_row:
        sum_row_start, sum_row_end = row, row
        for working_row in range(row + 1, max_row + 2):
            if (not is_blank_row(worksheet, working_row - 1)  # 上一行有值
                    and (ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value or is_blank_row(worksheet, working_row))):  # A列值不等于上一行 或 当前为空行（最后一次合并）
                # 求和
                sum_row_end = working_row - 1
                ws[f'E{sum_row_start}'] = range_sum(ws, f'C{sum_row_start}', f'C{sum_row_end}')
                # 合并居中
                ws[f'E{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws[f'A{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(f'E{sum_row_start}:E{sum_row_end}')
                ws.merge_cells(f'A{sum_row_start}:A{sum_row_end}')
                break
        row = sum_row_end + 1


def main():
    Current_Folder_path = os.getcwd()
    if not os.path.exists(Current_Folder_path + '\\New Folder'):
        os.mkdir(Current_Folder_path + '\\New Folder')
        # 遍历当前文件夹下的文件
    file_list = os.listdir(Current_Folder_path)
    docx_list = [a for a in file_list if a.endswith('.xlsx')]
    # 根据情况修改代码
    in_file_name = Current_Folder_path + '\\' + docx_list[0]
    processing_sheet = 'Sheet1'
    path_name = Current_Folder_path + '\\New Folder'
    out_file_name = 'Out.xlsx'
    wb = openpyxl.load_workbook(filename=os.path.join(path_name, in_file_name))
    total_amount(wb[processing_sheet])
    wb.save(os.path.join(path_name, out_file_name))


if __name__ == '__main__':
    main()
