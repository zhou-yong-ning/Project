import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from win32com import client as wc
import pandas as pd
import _thread
import threading
from multiprocessing import Process


def range_sum(worksheet, start, end):
    sum = 0
    for row in worksheet[start:end]:
        for cell in row:
            if cell.value != None:
                sum += cell.value
    return sum


def is_blank_row(worksheet, row_num, max_col=None):
    if not max_col:
        max_col = worksheet.max_column
    for cell in worksheet[row_num][:max_col]:
        if cell.value:
            return False
    return True


def total_amount(worksheet, fbfmc):
    """ 对某sheet的A、E列合并居中，并对E列求和 """
    ws = worksheet
    # 在第一行上方插入一行
    ws.insert_rows(1, 1)
    # 设置A1单元格值
    ws['A1'] = fbfmc + '地块属性表'
    # 设置字体样式
    fontsx = openpyxl.styles.Font(name=u'微软雅黑', bold=True, size=16)
    ws['A1'].font = fontsx
    # 设置A1单元格对齐方式
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    # 合并A1到G1单元格
    ws.merge_cells('A1:G1')
    # 指定第一行行高
    ws.row_dimensions[1].height = 30
    # 从第三行开始
    row = 3
    # 获取最大行
    max_row = ws.max_row
    ceel_range = ws['A2':f'M{max_row + 1}']
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            j.alignment = Alignment(horizontal="center", vertical="center")
    while row < max_row:
        sum_row_start, sum_row_end = row, row
        for working_row in range(row + 1, max_row + 2):
            if (not is_blank_row(worksheet, working_row - 1)  # 上一行有值
                    and (ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value)):  # A列值不等于上一行 或 当前为空行（最后一次合并）
                # 求和
                sum_row_end = working_row - 1
                # 求sum_row_start行到sum_row_end行之和
                # ws[f'D{working_row}'] = range_sum(ws, f'D{sum_row_start}', f'D{sum_row_end}')
                ws[f'E{sum_row_start}'] = range_sum(ws, f'D{sum_row_start}', f'D{sum_row_end}')
                # ws[f'F{working_row}'] = range_sum(ws, f'F{sum_row_start}', f'F{sum_row_end}')
                ws[f'G{sum_row_start}'] = range_sum(ws, f'F{sum_row_start}', f'F{sum_row_end}')
                # 合并居中
                # ws[f'B{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                # ws[f'A{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                # ws[f'E{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                # ws[f'G{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(f'B{sum_row_start}:B{sum_row_end}')
                ws.merge_cells(f'A{sum_row_start}:A{sum_row_end}')
                ws.merge_cells(f'E{sum_row_start}:E{sum_row_end}')
                ws.merge_cells(f'G{sum_row_start}:G{sum_row_end}')
                break
        row = sum_row_end + 1
    # 指定单元格值
    ws[f'C{max_row + 1}'] = '合计'
    # 设置单元格格式
    ws[f'C{max_row + 1}'].alignment = Alignment(horizontal="center", vertical="center")
    # 计算合计面积
    ws[f'E{max_row + 1}'] = range_sum(ws, f'D{3}', f'D{max_row}')
    ws[f'G{max_row + 1}'] = range_sum(ws, f'F{3}', f'F{max_row}')
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            j.alignment = Alignment(horizontal="center", vertical="center")
    # 设置列宽度
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 21
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 14


def dxc(dxclist, path):
    for file in dxclist:
        fbfmc = file.split('.')[0]
        # 根据情况修改代码
        processing_sheet = 'Sheet1'
        wb = openpyxl.load_workbook(filename=os.path.join(path, file))
        total_amount(wb[processing_sheet], fbfmc)
        wb.save(os.path.join(path, 'NewFolder', file))
        print(file + '调整完成')


if __name__ == '__main__':
    Current_Folder_path = os.getcwd()
    if not os.path.exists(Current_Folder_path + '\\NewFolder'):
        os.mkdir(Current_Folder_path + '\\NewFolder')
    # 遍历当前文件夹下的文件
    file_list0 = os.listdir(Current_Folder_path)
    # 提取后缀为xlsx的文件
    file_list = os.listdir(Current_Folder_path)
    xlsx_list = [a for a in file_list if a.endswith('.xlsx')]
    listnum = 50
    Arr = []
    for j in range(0, len(xlsx_list), listnum):
        if j + listnum <= len(xlsx_list):
            p = Process(target=dxc, args=(xlsx_list[j:j+listnum], Current_Folder_path))
            p.start()
            Arr.append(p)
        else:
            p = Process(target=dxc, args=(xlsx_list[j:len(xlsx_list) + 1], Current_Folder_path))
            p.start()
            Arr.append(p)
    print(Arr)
    for k in Arr:
        k.join()
