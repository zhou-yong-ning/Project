import pandas as pd
import os
# 自动调整列宽
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# 自动调整列宽方法
def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen * 1.2 + 2  # 也就是列宽为最大长度*1.2
    wb.save(filename)

if __name__ == '__main__':
    # col = input('请输入列标题: ')
    col = 'FBFBM'
    Current_Folder_path = os.getcwd()
    # 遍历当前文件夹下的文件
    file_list0 = os.listdir(Current_Folder_path)
    if not os.path.exists(Current_Folder_path + '\\NewFolder'):
        os.mkdir(os.path.join(Current_Folder_path, 'NewFolder'))
    file_list1 = [a for a in file_list0 if a.endswith('.xlsx')]
    print('正在读取文件')
    df = pd.read_excel(os.path.join(Current_Folder_path, file_list1[0]), dtype='object')
    df1 = pd.read_excel(os.path.join(Current_Folder_path, 'FBF.xlsx'), index_col='FBFBM', dtype='object')
    print('文件读取成功')
    # 获取某一列唯一值
    listType = df[col].unique()
    # 循环筛选导出
    for i in listType:
        df0 = df.loc[(df[col] == i)]
        # 这里需要转int类型，不然会报错，真不知道为什么，服了
        i = df1.loc[int(i), 'FBFMC']
        df0.to_excel(os.path.join(Current_Folder_path, 'NewFolder', i + '.xlsx'), index=False)
        # 自动调整列宽
        # reset_col(os.path.join(Current_Folder_path, 'NewFolder', str(i) + '.xlsx'))
        print(i + ' 表格导出成功')
