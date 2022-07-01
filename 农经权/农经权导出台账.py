import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from win32com import client as wc
import pandas as pd


def range_sum(worksheet, start, end):
    sum = 0
    for row in worksheet[start:end]:
        for cell in row:
            if cell.value != None:
                sum += cell.value
    return sum


def is_blank_row(worksheet, row_num, max_col=None):
    if not max_col:
        max_col = worksheet.max_column
    for cell in worksheet[row_num][:max_col]:
        if cell.value:
            return False
    return True


def total_amount(worksheet, fbfmc):
    """ 对某sheet的A、E列合并居中，并对E列求和 """
    ws = worksheet
    # 在第一行上方插入一行
    ws.insert_rows(1, 1)
    # 设置A1单元格值
    ws['A1'] = fbfmc + '地块属性表'
    # 设置字体样式
    fontsx = openpyxl.styles.Font(name=u'微软雅黑', bold=True, size=16)
    ws['A1'].font = fontsx
    # 设置A1单元格对齐方式
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    # 合并A1到G1单元格
    ws.merge_cells('A1:G1')
    # 指定第一行行高
    ws.row_dimensions[1].height = 30
    row = 3
    max_row = ws.max_row
    while row < max_row:
        sum_row_start, sum_row_end = row, row
        print(row, max_row)
        for working_row in range(row + 1, max_row + 2):
            if (not is_blank_row(worksheet, working_row - 1)  # 上一行有值
                    and (ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value)):  # A列值不等于上一行 或 当前为空行（最后一次合并）
                # 求和
                sum_row_end = working_row - 1
                # 求sum_row_start行到sum_row_end行之和
                # ws[f'D{working_row}'] = range_sum(ws, f'D{sum_row_start}', f'D{sum_row_end}')
                ws[f'E{sum_row_start}'] = range_sum(ws, f'D{sum_row_start}', f'D{sum_row_end}')
                # ws[f'F{working_row}'] = range_sum(ws, f'F{sum_row_start}', f'F{sum_row_end}')
                ws[f'G{sum_row_start}'] = range_sum(ws, f'F{sum_row_start}', f'F{sum_row_end}')
                # 合并居中
                ws[f'B{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws[f'A{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws[f'E{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws[f'G{sum_row_start}'].alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(f'B{sum_row_start}:B{sum_row_end}')
                ws.merge_cells(f'A{sum_row_start}:A{sum_row_end}')
                ws.merge_cells(f'E{sum_row_start}:E{sum_row_end}')
                ws.merge_cells(f'G{sum_row_start}:G{sum_row_end}')
                break
        row = sum_row_end + 1
    # 指定单元格值
    ws[f'C{max_row + 1}'] = '合计'
    # 设置单元格格式
    ws[f'C{max_row + 1}'].alignment = Alignment(horizontal="center", vertical="center")
    # 指定单元格合并
    ws.merge_cells(f'C{max_row + 1}:D{max_row + 1}')
    ws[f'E{max_row + 1}'] = range_sum(ws, f'D{3}', f'D{max_row}')
    ws[f'G{max_row + 1}'] = range_sum(ws, f'F{3}', f'F{max_row}')
    # 设置列宽度
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20



def main(Current_Folder_path, file):
    # 根据情况修改代码
    in_file_name = Current_Folder_path + '\\' + file
    processing_sheet = 'Sheet1'
    path_name = Current_Folder_path + '\\New Folder'
    out_file_name = file
    wb = openpyxl.load_workbook(filename=os.path.join(path_name, in_file_name))
    total_amount(wb[processing_sheet])
    wb.save(os.path.join(path_name, out_file_name))




if __name__ == '__main__':
    Current_Folder_path = os.getcwd()
    if not os.path.exists(Current_Folder_path + '\\New Folder'):
        os.mkdir(Current_Folder_path + '\\New Folder')
    # 遍历当前文件夹下的文件
    file_list0 = os.listdir(Current_Folder_path)
    # 提取后缀为xls的文件
    docx_list0 = [a for a in file_list0 if a.endswith('.xls')]
    file_list = os.listdir(Current_Folder_path)
    docx_list = [a for a in file_list if a.endswith('.xlsx')]
    # 将xls转换为xlsx
    if not len(docx_list0) == 0:
        try:
            word = wc.Dispatch("Excel.Application")
        except:
            word = wc.Dispatch("kwps.application")
        for A in docx_list0:
            doc = word.Workbooks.Open(Current_Folder_path + "\\" + A)
            A = A.replace('.docx', '.doc')
            doc.SaveAs(Current_Folder_path + "\\" + A + "x", 51)
            doc.Close()
            print(A + "转换完成")
            os.remove(Current_Folder_path + "\\" + A)
        word.Quit()
    for file in docx_list:
        df1 = pd.read_excel(Current_Folder_path + '\\' + file, dtype='object')

        fbfmc = file.split('.')[0]
        # df1.rename(columns={'DKBM': '地块编码',
        #                     'CBFBM': '承包方编码',
        #                     'YHTMJM': '原合同面积',
        #                     'HTMJM': '发证面积',
        #                     'CBFMC': '承包方名称'}, inplace=True)
        # df1['原合同面积合计'] = ''
        # df1['发证面积合计'] = ''
        # df1 = df1.loc[:, ['承包方编码', '承包方名称', '地块编码', '原合同面积', '原合同面积合计', '发证面积', '发证面积合计']]
        print(fbfmc)
        # 根据'CBFBM'进行降序排列
        df1.sort_values(by=['承包方编码'], ascending=True, inplace=True)
        df1.to_excel(Current_Folder_path + '\\' + file, index=False)

        # 根据情况修改代码
        in_file_name = Current_Folder_path + '\\' + file
        processing_sheet = 'Sheet1'
        path_name = Current_Folder_path + '\\New Folder'
        out_file_name = file
        wb = openpyxl.load_workbook(filename=os.path.join(path_name, in_file_name))
        total_amount(wb[processing_sheet], fbfmc)
        wb.save(os.path.join(path_name, out_file_name))
