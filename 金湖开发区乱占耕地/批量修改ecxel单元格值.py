import openpyxl
import os


def xlsx_value(file,str_txt):
    # 读取文件
    workbook = openpyxl.load_workbook(file)
    # 指定当前工作表
    worksheet = workbook.active
    # 读取单元格值
    # A1_value = worksheet.cell(1, 1).value
    # 写入单元格值
    worksheet.cell(1, 1, value=str_txt)
    # 保存
    workbook.save(file)


Current_Folder_path = os.getcwd()
# 提取后缀为xlsx的文件
file_list = os.listdir(Current_Folder_path)
file_list = [a for a in file_list if a.endswith('.xlsx')]
for i in file_list:
    file_name = i.split('.')[0]
    txt = ''.join(['大冶市金湖街道办事处',file_name,'一户一宅”情况明细表'])
    xlsx_value(i,txt)
