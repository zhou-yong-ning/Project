import pandas as pd
from win32com import client as wc
import os
# 自动调整列宽
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen * 1.1 + 0  # 也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)
# 以上为自动调整列宽方法
# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xls的文件
docx_list = [a for a in file_list if a.endswith('.xls')]
# 将xls转换为xlsx
if not len(docx_list) == 0:
    try:
        word = wc.Dispatch("Excel.Application")
    except:
        word = wc.Dispatch("kwps.application")
    for A in docx_list:
        doc = word.Workbooks.Open(Current_Folder_path + "\\" + A)
        A = A.replace('.docx', '.doc')
        doc.SaveAs(Current_Folder_path + "\\" + A + "x", 51)
        doc.Close()
        print(A + "转换完成")
        os.remove(Current_Folder_path + "\\" + A)
    word.Quit()
# 提取文件内容
df0 = pd.read_excel(Current_Folder_path + "\\" + "宗地属性表.xlsx", dtype='object')
df1 = pd.read_excel(Current_Folder_path + "\\" + "房地一体.xlsx", dtype='object')
# 提取行索引所对应的内容 宗地码    土地坐落  实测面积    东     南     西     北     权利人名称 发证情况
df0 = df0.loc[:, ['ZDTYBM', 'TDZL', 'SCMJ', 'SZD', 'SZN', 'SZX', 'SZB', 'QLRMC', 'FZZT']]
df1 = df1.loc[:, ['宗地代码', '证件号码', '关系']]
# 筛选
df1 = df1.loc[(df1['关系'] == '户主')]
df1 = df1.loc[:, ['宗地代码', '证件号码']]
# 列索引重命名
df1.rename(columns={'宗地代码': 'ZDTYBM','证件号码':'身份证号码'}, inplace=True)
# 暴力重命名
# df1.columns = ['ZDTYBM','证件号码']
# merge函数匹配
df1to0 = pd.merge(df0, df1, on='ZDTYBM', how='left')
df0 = df1to0
df1 = pd.read_excel(Current_Folder_path + "\\" + "成员信息.xlsx", dtype='object')
mingchen = df1.loc[0,'Unnamed: 1']
daima = df1.loc[0,'Unnamed: 3']
df1 = pd.read_excel(Current_Folder_path + "\\" + "成员信息.xlsx", dtype='object',header=4)
df1 = df1.loc[:,['Unnamed: 6','户主联系电话']]
# 列索引重命名
df1.rename(columns={'Unnamed: 6': '身份证号码','Unnamed: 13':'电话号码'}, inplace=True)
df1to0 = pd.merge(df0, df1, on='身份证号码', how='left')
df0 = df1to0
# df1 = pd.read_excel(Current_Folder_path + "\\" + "成员信息.xlsx", dtype='object')
# 新增一列固定值
df0['证件类型'] = '居民身份证'
df0['宅基地代码'] = daima
df0['所有权性质'] = '集体土地所有权'
df0['区域代码'] = '420281'
df0['代表人证件类型'] = '居民身份证'
df0['邮政编码'] = '435100'
df0['是否成立农村集体经济组织'] = '是'
df0['农村集体经济组织代码'] = daima
df0['农村集体经济组织名称'] = mingchen
df0['发包方代码'] = daima
df0['发包方名称'] = mingchen
# 复制并在首列创建新列
df_1 = df0.TDZL
df0.insert(1,'通讯地质',df_1)
df0.insert(1,'代表人地址',df_1)
df0.rename(columns={'SZD':'宗地四至（东）',
                    'SZN':'宗地四至（南）',
                    'SZX':'宗地四至（西）',
                    'SZB':'宗地四至（北）',
                    'QLRMC':'权利人名称',
                    'FZZT':'是否发证',
                    'TDZL':'坐落',
                    'SCMJ':'宗地面积'},inplace=True)
# if 'a' in df0.columns:
#     print('zai')
# else:
#     print('fou')
df0.to_excel(Current_Folder_path + '\\汇总表.xlsx', index=False)
# 调用调整列宽方法
reset_col(Current_Folder_path + '\\汇总表.xlsx')

