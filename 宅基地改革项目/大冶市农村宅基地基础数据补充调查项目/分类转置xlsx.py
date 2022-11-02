import os
import openpyxl
from openpyxl.styles import Alignment

def total_amount(worksheet):
    ws = worksheet
    # 从第二行开始
    row = 2
    # 获取最大行
    max_row = ws.max_row
    while row < max_row:
        sum_row_start, sum_row_end = row, row
        for working_row in range(row + 1, max_row + 2):
            if ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value:  # A列值不等于上一行或当前为空行
                for i in range(2,working_row-row+2):
                    sum_row_end = working_row - 1
                    # 根据列的数字返回字母
                    j = openpyxl.utils.get_column_letter(i)
                    # 为指定单元格赋值
                    ws[f'{j}{row}'] = ws[f'B{row+i-2}'].value
                    print(j,row,ws[f'B{row+i-2}'].value,row)
                ws.delete_rows(row+1, sum_row_end-row)  # 其中1指删除的位置，2指删除的行数
                break
        row = row+1
        max_row = ws.max_row
        # 循环调整单元格格式
    ceel_range = ws['A2':f'Z{max_row + 1}']
    for i in ceel_range:
        for j in i:
            # 设置居中对齐（上下居中，左右居中）
            j.alignment = Alignment(horizontal="center", vertical="center")



if __name__ == '__main__':
    Current_Folder_path = os.getcwd()
    # 遍历当前文件夹下的文件
    file_list = os.listdir(Current_Folder_path)
    xlsx_list = [a for a in file_list if a.endswith('.xlsx')]
    xlsx = xlsx_list[0]
    wb = openpyxl.load_workbook(filename=os.path.join(Current_Folder_path, xlsx))
    # 获取当前活跃的工作表
    worksheet = wb.active
    total_amount(worksheet)
    xlsx = xlsx.split('.')[0] + '(转置)' + '.xlsx'
    wb.save(os.path.join(Current_Folder_path,xlsx))
    print('调整完成')
