import pandas as pd
import os
from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import warnings


def jiequ1(x):
    x = str(x)
    return x[-5:]


def jiequ2(x):
    x = str(x)
    return x[:19]


def jiequ3(x):
    x = str(x)
    return x[12:19]


# 宗地权利人农户问题汇总表
def total_amount(worksheet):
    ws = worksheet
    # 获取最大行
    max_row = ws.max_row
    # 指定单元格范围
    ceel_range = ws['A2':f'Q{max_row + 1}']
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            # 设置居中对齐（上下居中，左右居中）
            j.alignment = Alignment(horizontal="center", vertical="center")
    # 指定列宽列表A   B   C   D   E   F   G   H I   J  K   L   M   N   Q   P
    listnum = [25, 20, 20, 10, 10, 20, 20, 9, 9, 12, 20, 5, 26, 22, 40, 30, 30]
    for j in range(1, len(listnum) + 1):
        # 根据列的数字返回字母
        k = openpyxl.utils.get_column_letter(j)
        # 根据listnum列表设置列宽
        ws.column_dimensions[k].width = listnum[j - 1]
    row = 2
    print('以下为多户一宅宅基地')
    # 创建多户一宅列表
    DHYZ_list=[]
    while row < max_row:
        sum_row_start, sum_row_end = row, row
        for working_row in range(row + 1, max_row + 2):
            if ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value:  # A列值不等于上一行 或 当前为空行（最后一次合并）
                sum_row_end = working_row - 1
                # 合并
                ws.merge_cells(f'B{sum_row_start}:B{sum_row_end}')
                ws.merge_cells(f'A{sum_row_start}:A{sum_row_end}')
                # 多户一宅内农户合并
                if ws[f'P{sum_row_start}'].value != ws[f'P{sum_row_end}'].value:
                    print(ws[f'A{sum_row_start}'].value)
                    DHYZ_list.append(ws[f'A{sum_row_start}'].value)
                    row1 = sum_row_start
                    while row1 < sum_row_end:
                        sum_row_start1, sum_row_end1 = row1, row1
                        for working_row1 in range(row1 + 1, sum_row_end+2):
                            if ws[f'P{working_row1}'].value != ws[f'P{working_row1 - 1}'].value \
                                    or working_row1-1 == sum_row_end:  # P列值不等于上一行 或 当前为最后一行
                                sum_row_end1 = working_row1-1
                                ws.merge_cells(f'C{sum_row_start1}:C{sum_row_end1}')
                                ws.merge_cells(f'D{sum_row_start1}:D{sum_row_end1}')
                                ws.merge_cells(f'E{sum_row_start1}:E{sum_row_end1}')
                                ws.merge_cells(f'F{sum_row_start1}:F{sum_row_end1}')
                                ws.merge_cells(f'G{sum_row_start1}:G{sum_row_end1}')
                                break
                        row1 = sum_row_end1 + 1
                else:
                    ws.merge_cells(f'C{sum_row_start}:C{sum_row_end}')
                    ws.merge_cells(f'D{sum_row_start}:D{sum_row_end}')
                    ws.merge_cells(f'E{sum_row_start}:E{sum_row_end}')
                    ws.merge_cells(f'F{sum_row_start}:F{sum_row_end}')
                    ws.merge_cells(f'G{sum_row_start}:G{sum_row_end}')
                break
        row = sum_row_end + 1
    print('以上为多户一宅宅基地')
    return DHYZ_list
# 宗地房屋问题汇总表
def total_amount1(worksheet):
    ws = worksheet
    # 获取最大行
    max_row = ws.max_row
    ceel_range = ws['A2':f'Q{max_row + 1}']
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            j.alignment = Alignment(horizontal="center", vertical="center")
    listnum = [25, 10, 10, 29, 29, 18, 13, 8, 9, 12, 12, 15, 12, 20, 15, 30]
    for j in range(1, len(listnum) + 1):
        # 根据列的数字返回字母
        k = openpyxl.utils.get_column_letter(j)
        # 根据listnum列表设置列宽
        ws.column_dimensions[k].width = listnum[j - 1]
    row = 2
    while row < max_row:
        sum_row_start, sum_row_end = row, row
        for working_row in range(row + 1, max_row + 2):
            if ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value:  # A列值不等于上一行 或 当前为空行（最后一次合并）
                sum_row_end = working_row - 1
                # 合并
                ws.merge_cells(f'D{sum_row_start}:D{sum_row_end}')
                ws.merge_cells(f'A{sum_row_start}:A{sum_row_end}')
                ws.merge_cells(f'C{sum_row_start}:C{sum_row_end}')
                ws.merge_cells(f'L{sum_row_start}:L{sum_row_end}')
                ws.merge_cells(f'O{sum_row_start}:O{sum_row_end}')
                break
        row = sum_row_end + 1

# 忽略警告
warnings.filterwarnings('ignore')
# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xlsx的文件
docx_list = [a for a in file_list if a.endswith('.xlsx')]
# 利用推导式获取指定表数据
syqrxx = [i for i in docx_list if '使用权人表数据' in i]
fwxx = [i for i in docx_list if '房屋信息表数据' in i]
zdxx = [i for i in docx_list if '宗地信息表数据' in i]
cyxx = [i for i in docx_list if '户内成员表数据' in i]
# 设置保存路径
save_path = os.path.join(Current_Folder_path, '导出汇总表.xlsx')
# 以下部分制作宗地权利人农户问题汇总表
# 提取文件内容，
df0 = pd.read_excel(os.path.join(Current_Folder_path, syqrxx[0]), dtype='object')
print('使用权人表读取成功')
df2 = pd.read_excel(os.path.join(Current_Folder_path, zdxx[0]), dtype='object')
print('宗地信息表读取成功')
df4 = pd.read_excel(os.path.join(Current_Folder_path, cyxx[0]), dtype='object')
print('户内成员表读取成功')
# 提取行标题所对应的内容
df0 = df0.loc[:,
      ['宅基地代码', '使用权人代表姓名', '使用权人代表证件类型',
       '权利人类型', '国家/地区', '是否使用权人之间共有', '农户代码']]
# 提取行索引所对应的内容
df2 = df2.loc[:, ['宅基地代码', '宗地图预编码']]
# 提取户内成员信息
df4 = df4.loc[:,
      ['农户代码', '姓名', '证件号码', '户口类型', '性别', '与户主关系', '数据来源', '是否本集体经济组织成员',
       '通讯地址']]
# 列索引重命名
df4.rename(columns={'姓名': '成员姓名'}, inplace=True)
# 通过房屋表与宗地表连接宅基地代码与宗地号
df02 = pd.merge(df2, df0, on='宅基地代码', how='left')
# 通过农户代码连接家庭成员信息
df402 = pd.merge(df02, df4, on='农户代码', how='left')
print('宗地权利人农户问题汇总表数据连接成功')
# 调整列顺序
df4012 = df402[
    ['宅基地代码', '宗地图预编码', '使用权人代表姓名',
     '权利人类型', '国家/地区', '是否使用权人之间共有', '使用权人代表证件类型',
     '户口类型', '成员姓名', '与户主关系', '证件号码', '性别', '数据来源', '是否本集体经济组织成员', '通讯地址',
     '农户代码']]
# 保存多个sheet在同一工作簿中，需要用到函数pd.ExcelWriter()。
write = pd.ExcelWriter(save_path)
# 保存至指定工作表
df4012.to_excel(write, sheet_name='宗地权利人农户问题汇总表', index=False)
write.save()
print('宗地权利人农户问题汇总表完成')

# 提取文件内容至宗地房屋问题汇总表
df0 = pd.read_excel(os.path.join(Current_Folder_path, fwxx[0]), dtype='object')
print('房屋信息表读取成功')
df1 = pd.read_excel(os.path.join(Current_Folder_path, zdxx[0]), dtype='object')
print('宗地信息表读取成功')
# 提取行索引所对应的内容
df0 = df0.loc[:, ['宅基地代码', '坐落', '房屋结构', '数据来源', '备注', '房屋状态',
                  '利用状况', '房屋是否闲置', '实际层数', '农民房屋面积', '农民房屋代码', '自然栋号']]
df1 = df1.loc[:, ['宅基地代码', '数据来源', '宅基地利用状况', '历史遗留问题','宗地图预编码']]
# 列索引重命名
df0.rename(columns={'数据来源': '幢数据来源', '利用状况': '房屋利用状况'}, inplace=True)
df0['房屋幢号'] = df0['自然栋号'].apply(jiequ1)
df1['宗地号'] = df1['宗地图预编码'].apply(jiequ3)
df1.rename(columns={'数据来源': '宅基地数据来源'}, inplace=True)
# merge函数匹配
df1to0 = pd.merge(df1, df0, on='宅基地代码', how='left')
# 提取指定列数据并按顺序排列
df1to0 = df1to0[['宅基地代码', '房屋幢号', '宗地号', '宅基地数据来源', '幢数据来源', '备注', '农民房屋面积', '实际层数',
                 '房屋状态', '房屋是否闲置', '房屋利用状况', '宅基地利用状况', '房屋结构', '坐落', '历史遗留问题',
                 '农民房屋代码', '自然栋号']]
# 自定义排序
df1to0.sort_values(by=['宅基地代码', '房屋幢号'], inplace=True, ascending=True)
print('宗地房屋问题汇总表数据连接成功')
df1to0.to_excel(write, sheet_name='宗地房屋问题汇总表', index=False)
write.save()
# write必需关闭，否则无法用openpyxl打开工作表
write.close()
print('宗地房屋问题汇总表完成')
wb = openpyxl.load_workbook(filename=save_path)
total_amount1(wb['宗地房屋问题汇总表'])
print('宗地房屋问题汇总表格式调整完成')
dhyz = total_amount(wb['宗地权利人农户问题汇总表'])
print('宗地权利人农户问题汇总表格式调整完成')
wb.save(save_path)
wb.close()
# 将多户一宅列表写入工作表
dhyz.insert(0,'多户一宅宅基地代码')
file_path = os.path.join(Current_Folder_path,'导出汇总表.xlsx')
workbook = openpyxl.load_workbook(file_path)
# 新建工作表
workbook.create_sheet('多户一宅')
# 指定工作表
worksheet = workbook['多户一宅']
# 写入列表数据
r = 1
for i in dhyz:
    worksheet.cell(row=r, column=1).value = i
    r += 1
workbook.save(file_path)

