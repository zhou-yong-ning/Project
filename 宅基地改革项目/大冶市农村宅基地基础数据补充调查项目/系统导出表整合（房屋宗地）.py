import pandas as pd
import os
from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def jiequ1(x):
    x = str(x)
    return x[-5:]


def jiequ2(x):
    x = str(x)
    return x[12:19]


def total_amount(worksheet):
    ws = worksheet
    # 获取最大行
    max_row = ws.max_row
    ceel_range = ws['A2':f'Q{max_row + 1}']
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            j.alignment = Alignment(horizontal="center", vertical="center")
    listnum = [25, 10, 10, 29, 29, 18, 13, 8, 9, 12, 12, 15, 12, 20, 15, 30]
    for j in range(1, len(listnum)+1):
        # 根据列的数字返回字母
        k = openpyxl.utils.get_column_letter(j)
        # 根据listnum列表设置列宽
        ws.column_dimensions[k].width = listnum[j - 1]
    row = 2
    while row < max_row:
        sum_row_start, sum_row_end = row, row
        for working_row in range(row + 1, max_row + 2):
            if ws[f'A{working_row}'].value != ws[f'A{working_row - 1}'].value:  # A列值不等于上一行 或 当前为空行（最后一次合并）
                # 求和
                sum_row_end = working_row - 1
                # 合并
                ws.merge_cells(f'D{sum_row_start}:D{sum_row_end}')
                ws.merge_cells(f'A{sum_row_start}:A{sum_row_end}')
                ws.merge_cells(f'C{sum_row_start}:C{sum_row_end}')
                ws.merge_cells(f'L{sum_row_start}:L{sum_row_end}')
                ws.merge_cells(f'O{sum_row_start}:O{sum_row_end}')
                break
        row = sum_row_end + 1

# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xls的文件
docx_list = [a for a in file_list if a.endswith('.xlsx')]
# 利用推导式获取“房屋信息表”
fwxx = [i for i in docx_list if '房屋信息' in i]
zdxx = [i for i in docx_list if '宗地信息' in i]
# # 提取文件内容
df0 = pd.read_excel(os.path.join(Current_Folder_path, fwxx[0]), dtype='object')
print('房屋信息表数据读取成功')
df1 = pd.read_excel(os.path.join(Current_Folder_path, zdxx[0]), dtype='object')
print('宗地信息表读取成功')
# 提取行索引所对应的内容
df0 = df0.loc[:, ['宅基地代码', '坐落', '房屋结构', '数据来源', '备注', '房屋状态',
                  '利用状况', '房屋是否闲置', '实际层数', '农民房屋面积', '农民房屋代码', '自然栋号']]
df1 = df1.loc[:, ['宅基地代码', '数据来源', '宅基地利用状况','历史遗留问题']]
# 列索引重命名
df0.rename(columns={'数据来源': '幢数据来源', '利用状况': '房屋利用状况'}, inplace=True)
df0['房屋幢号'] = df0['自然栋号'].apply(jiequ1)
df0['宗地号'] = df0['自然栋号'].apply(jiequ2)
df1.rename(columns={'数据来源': '宅基地数据来源'}, inplace=True)
# merge函数匹配
df1to0 = pd.merge(df0, df1, on='宅基地代码', how='left')
df1to0 = df1to0[['宅基地代码', '房屋幢号', '宗地号', '宅基地数据来源', '幢数据来源', '备注', '农民房屋面积', '实际层数',
                 '房屋状态', '房屋是否闲置', '房屋利用状况', '宅基地利用状况', '房屋结构', '坐落', '历史遗留问题','农民房屋代码', '自然栋号']]
# 自定义排序
df1to0.sort_values(by=['宅基地代码','房屋幢号'], inplace=True, ascending=True)
print('数据连接成功')
sava_path = os.path.join(Current_Folder_path, "房屋宗地问题汇总表.xlsx")
df1to0.to_excel(sava_path, index=False)
print('数据写入成功')
wb = openpyxl.load_workbook(filename=sava_path)
total_amount(wb['Sheet1'])
wb.save(sava_path)
print('格式调整完成')
