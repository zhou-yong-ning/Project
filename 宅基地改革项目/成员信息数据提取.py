import pandas as pd
from win32com import client as wc
import os
# 自动调整列宽
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen * 1.2 + 2  # 也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)


# 以上为自动调整列宽方法

def huzhu(guanxi):
    if guanxi == '户主':
        return "是"
    else:
        return "否"


# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
if not os.path.exists(Current_Folder_path + '\\New Folder'):
    os.mkdir(Current_Folder_path + '\\New Folder')
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xls的文件
docx_list = [a for a in file_list if a.endswith('.xls')]
# 将xls转换为xlsx
if not len(docx_list) == 0:
    try:
        word = wc.Dispatch("Excel.Application")
    except:
        word = wc.Dispatch("kwps.application")
    for A in docx_list:
        doc = word.Workbooks.Open(Current_Folder_path + "\\" + A)
        A = A.replace('.docx', '.doc')
        doc.SaveAs(Current_Folder_path + "\\" + A + "x", 51)
        doc.Close()
        print(A + "转换完成")
        os.remove(Current_Folder_path + "\\" + A)
    word.Quit()

docx_list = [a for a in file_list if a.endswith('.xlsx')]
# 利用推导式获取“.*成员信息.xlsx”数据表
cyxx = [i for i in docx_list if i[-9:] == '成员信息.xlsx']
print(cyxx[0])
# 提取文件内容
df0 = pd.read_excel(Current_Folder_path + "\\" + docx_list[0], dtype='object')
mingchen = df0.loc[0, 'Unnamed: 1']
daima = df0.loc[0, 'Unnamed: 3']
# 删除指定行
df0 = df0.drop(df0.index[[0, 1, 2, 3]])
df0 = df0.loc[:, ['Unnamed: 1', 'Unnamed: 3',
                  'Unnamed: 4', 'Unnamed: 6', 'Unnamed: 7',
                  'Unnamed: 12', 'Unnamed: 13']]
df0.rename(columns={'Unnamed: 1': '农户预编码',
                    'Unnamed: 3': '姓名',
                    'Unnamed: 4': '性别',
                    'Unnamed: 6': '证件号码',
                    'Unnamed: 7': '与户主关系',
                    'Unnamed: 12': '通讯地址',
                    'Unnamed: 13': '联系电话',
                    }, inplace=True)
# print(df0)
# 替换某一列信息并强制转换值类型
df0.loc[:, "与户主关系"] = df0["与户主关系"].str.replace("本人", "户主").astype('str')
df0['是否户主'] = df0.apply(lambda x: huzhu(x.与户主关系), axis=1)
df0['是否五保户'] = '否'
df0['是否贫困户'] = '否'
df0['是否本集体经济组织成员'] = '是'
df0['资格权保障情况'] = ''
df0['证件类型'] = '身份证'
df0['户口类型'] = ''
df0['户籍是否在本村'] = '是'
df0['户籍不在本村原因'] = ''
df0['户籍不在本村原因其他'] = ''
df0['成员备注'] = ''
df0['成员备注说明'] = ''
df0['备注'] = ''
print(docx_list[0])
csx = input('请输入村名拼音简写： ')
# 跟据现有信息新增列
df0['农户预编码'] = df0.apply(lambda x: csx + x.农户预编码, axis=1)
df0 = df0[['农户预编码', '通讯地址', '是否户主',
           '是否五保户', '是否贫困户', '资格权保障情况',
           '姓名', '证件类型', '证件号码', '性别', '与户主关系',
           '户口类型', '联系电话', '户籍是否在本村', '户籍不在本村原因', '户籍不在本村原因其他',
           '是否本集体经济组织成员', '成员备注', '成员备注说明', '备注']]
# 创建不同的工作表
write = pd.ExcelWriter(Current_Folder_path + '\\New Folder\\' + docx_list[0])
df0.to_excel(write, sheet_name='农户及户内成员', index=False)
write.save()
df1 = df0.loc[:, ['姓名', '证件类型', '证件号码', '与户主关系']]
df1 = df1.loc[(df1['与户主关系'] == '户主')]
df1 = df1.drop(columns='与户主关系')
df1['所有权人代码'] = ''
df1.loc[4, '所有权人代码'] = daima
df1['宗地预编码'] = ''
df1['农民房屋预编码'] = ''
df1['农户预编码'] = ''
df1['使用权人代码'] = ''
df1['不动产单元号'] = ''
df1['不动产权证号'] = ''
df1['权证印刷序列号'] = ''
df1['发证机关'] = ''
df1['所属行业'] = ''
df1['国家/地区'] = ''
df1['户籍所在省市'] = ''
df1['性别'] = ''
df1['电话'] = ''
df1['地址'] = ''
df1['是否使用权人之间共有'] = ''
df1['分摊宗地面积'] = ''
df1['权利人类型'] = ''
df1['是否本农村集体经济组织成员'] = ''
df1['户口类型'] = ''
df1['备注'] = ''
df1['区域代码'] = ''
df1.to_excel(write, sheet_name='使用权人', index=False)
write.save()
df2 = pd.DataFrame([['所有权人代码', '所有权人名称', '所有权性质', '所有权确权登记面积', '代表人姓名',
                     '代表人证件类型','代表人证件号码','代表人联系电话','代表人通讯地址','代表人邮政编码',
                     '代理人姓名','代理人证件类型','代理人证件号码','代理人联系电话','代理人通讯地址',
                     '代理人邮政编码','是否成立农村集体经济组织','区域代码','发包方代码']])
df2.to_excel(write,sheet_name='所有权人',header=False,index=False)
write.close()
reset_col(Current_Folder_path + '\\New Folder\\' + docx_list[0])
print('提取成功')
