import pandas as pd
import os
from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def jiequ1(x):
    x = str(x)
    return x[-5:]


def jiequ2(x):
    x = str(x)
    return x[:19]


def total_amount(worksheet):
    ws = worksheet
    # 获取最大行
    max_row = ws.max_row
    # 指定单元格范围
    ceel_range = ws['A2':f'M{max_row + 1}']
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            # 设置居中对齐（上下居中，左右居中）
            j.alignment = Alignment(horizontal="center", vertical="center")
    # 指定列宽列表1   2   3   4   5   6   7   8  9  10  11  12  13  14  15  16
    listnum = [25, 20, 20, 25, 16, 19, 22, 9, 26, 12, 12, 15, 12, 20, 30, 30]
    for j in range(1, 14):
        # 根据列的数字返回字母
        k = openpyxl.utils.get_column_letter(j)
        # 根据listnum列表设置列宽
        ws.column_dimensions[k].width = listnum[j - 1]


# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xls的文件
docx_list = [a for a in file_list if a.endswith('.xlsx')]
# 利用推导式获取“房屋信息表”
syqrxx = [i for i in docx_list if '使用权人表数据' in i]
fwxx = [i for i in docx_list if '房屋信息' in i]
zdxx = [i for i in docx_list if '宗地信息' in i]
# cyxx = [i for i in docx_list if '户内成员' in i]
# # 提取文件内容
df0 = pd.read_excel(os.path.join(Current_Folder_path, syqrxx[0]), dtype='object')
print('使用权人表数据读取成功')
df1 = pd.read_excel(os.path.join(Current_Folder_path, fwxx[0]), dtype='object')
print('房屋信息表读取成功')
df2 = pd.read_excel(os.path.join(Current_Folder_path, zdxx[0]), dtype='object')
print('宗地信息表读取成功')
df3 = pd.read_excel(
    r"E:\WORK\院宅基地改革项目\宅基地改革\资料收集\房地一体数据20220610\权利人20220610（参考）\金山店镇.xlsx",
    dtype='object')
# 筛选提取户主关系列为‘户主’或者‘本人’的数据表
# df3 = df3[(df3['关系']=='户主') | (df3['关系']=='本人')]
# 删除重复值(保留第一个)
df3.drop_duplicates(subset='宗地代码', inplace=True, keep='first')
# 提取行索引所对应的内容
df0 = df0.loc[:,
      ['宅基地代码', '使用权人代表姓名', '使用权人代表证件类型', '使用权人代表证件号码', '是否本农村集体经济组织成员',
       '户口类型', '权利人类型', '国家/地区', '是否使用权人之间共有', '农户代码']]
# 截取宗地号
df1['宗地号'] = df1['自然栋号'].apply(jiequ2)
# 提取行索引所对应的内容
df1 = df1.loc[:, ['宅基地代码', '宗地号']]
# 排序
df1.sort_values(by='宗地号', inplace=True, ascending=True)
# 删除重复值(保留第一个)
df1.drop_duplicates(subset='宅基地代码', inplace=True, keep='first')
# 提取行索引所对应的内容
df2 = df2.loc[:, '宅基地代码']
# 提取行索引所对应的内容
df3 = df3.loc[:, ['宗地代码', '图形权利人名称']]
# 列索引重命名
df3.columns = ['宗地号', '房地一体权利人名称']
# 通过房屋表与宗地表连接宅基地代码与宗地号
df12 = pd.merge(df2, df1, on='宅基地代码', how='left')
print('数据连接成功（1）')
# 通过宗地号连接原房地一体数据
df312 = pd.merge(df12, df3, on='宗地号', how='left')
print('数据连接成功（2）')
# 通过宅基地代码连接使用权人表
df0312 = pd.merge(df312, df0, on='宅基地代码', how='left')
# 调整列顺序
df0312 = df0312[
    ['宅基地代码', '宗地号', '农户代码', '房地一体权利人名称', '使用权人代表姓名', '是否本农村集体经济组织成员',
     '权利人类型', '国家/地区','是否使用权人之间共有', '使用权人代表证件类型', '使用权人代表证件号码', '户口类型']]
sava_path = os.path.join(Current_Folder_path, "宗地权利人问题汇总表.xlsx")
df0312.to_excel(sava_path, index=False)
print('数据写入成功')
wb = openpyxl.load_workbook(filename=sava_path)
total_amount(wb['Sheet1'])
wb.save(sava_path)
print('格式调整完成')
