import pandas as pd
import os
# 自动调整列宽
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import openpyxl


def total_amount(worksheet):
    ws = worksheet
    # 获取最大行
    max_row = ws.max_row
    ceel_range = ws['A2':f'M{max_row + 1}']
    # 循环调整单元格格式
    for i in ceel_range:
        for j in i:
            j.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 29
    ws.column_dimensions['C'].width = 29
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 30
    ws.column_dimensions['M'].width = 30


# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xls的文件
docx_list = [a for a in file_list if a.endswith('.xlsx')]
# 利用推导式获取“房屋信息表”
fwxx = [i for i in docx_list if '房屋信息' in i]
zdxx = [i for i in docx_list if '宗地信息' in i]
# # 提取文件内容
df0 = pd.read_excel(os.path.join(Current_Folder_path, fwxx[0]), dtype='object')
print('房屋信息表数据读取成功')
df1 = pd.read_excel(os.path.join(Current_Folder_path, zdxx[0]), dtype='object')
print('宗地信息表读取成功')
# 提取行索引所对应的内容
df0 = df0.loc[:, ['宅基地代码', '坐落', '房屋结构', '数据来源', '备注', '房屋状态', '利用状况', '房屋是否闲置', '实际层数', '农民房屋面积', '农民房屋代码']]
df1 = df1.loc[:, ['宅基地代码', '数据来源', '宅基地利用状况']]
# 列索引重命名
df0.rename(columns={'数据来源': '幢数据来源', '利用状况': '房屋利用状况'}, inplace=True)
df1.rename(columns={'数据来源': '宅基地数据来源'}, inplace=True)
# merge函数匹配
df1to0 = pd.merge(df0, df1, on='宅基地代码', how='left')
df1to0 = df1to0[['宅基地代码', '宅基地数据来源', '幢数据来源', '备注', '农民房屋面积', '实际层数',
                 '房屋状态', '房屋是否闲置', '房屋利用状况', '宅基地利用状况', '房屋结构', '坐落', '农民房屋代码']]
# 排序
df1to0.sort_values(by='宅基地代码', inplace=True, ascending=False)
print('数据连接成功')
sava_path = os.path.join(Current_Folder_path, "问题汇总表.xlsx")
df1to0.to_excel(sava_path, index=False)
print('数据写入成功')
wb = openpyxl.load_workbook(filename=sava_path)
total_amount(wb['Sheet1'])
wb.save(sava_path)
print('格式调整完成')
